"""Build a blind forced-choice (A/B) preference workbook to compare two runs.

Implements the MT-evaluation expert's pairwise design: present the reference
together with two anonymized system outputs in randomized order and let a judge
pick which is *closest* to the reference (ties allowed). This resolves model
selection when absolute metrics are statistically tied (e.g. v2.1 vs v2.1b).

Default direction is shw2spa, because the outputs are in Spanish and therefore
judgeable without a Shiwilu speaker. The per-row A/B mapping is written to a
separate anon-key JSON so the workbook itself stays blind.

Output:
    reports/05_nmt/evaluation_xl/pairwise_preference.xlsx
    reports/05_nmt/evaluation_xl/pairwise_preference_anon_key.json
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import random
import re
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from scripts.nmt._paths import resolve_paths  # noqa: E402

DEFAULT_RUN_A = "nllb_bidi_lora_v2_1b_loraplus_xl"
DEFAULT_RUN_B = "nllb_bidi_lora_v2_1_dora_loraplus_xl"

SPEAKER_COLUMNS = [
    "N",
    "Texto fuente",
    "Referencia",
    "Opción A",
    "Opción B",
    "Más cercana (A/B/Empate)",
    "Comentario",
]


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--variant", choices=["main", "xl"], default="xl")
    p.add_argument("--run-a", default=DEFAULT_RUN_A, help="First system (default: champion v2.1b).")
    p.add_argument("--run-b", default=DEFAULT_RUN_B, help="Second system (default: v2.1 DoRA+LoRA+).")
    p.add_argument(
        "--directions",
        nargs="+",
        default=["shw2spa"],
        choices=["shw2spa", "spa2shw"],
        help="Directions to include (default: shw2spa, the Spanish-output side).",
    )
    p.add_argument("--per-direction", type=int, default=60)
    p.add_argument("--seed", type=int, default=2026)
    p.add_argument("--output", type=Path, default=None)
    p.add_argument("--anon-key", type=Path, default=None)
    return p.parse_args()


def _read_jsonl(path: Path) -> list[dict]:
    rows: list[dict] = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def _display_path(path: Path) -> str:
    path = path.resolve()
    try:
        return str(path.relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path)


def _normalized(text: str) -> str:
    return re.sub(r"\s+", " ", str(text).strip()).casefold()


def _has_broken_encoding(text: str) -> bool:
    markers = ("�", "Ã", "Â", "â€™", "â€œ", "â€")
    return any(marker in str(text) for marker in markers)


def _usable(*texts: str) -> bool:
    if any(not str(t).strip() for t in texts):
        return False
    return not any(_has_broken_encoding(str(t)) for t in texts)


def _index_by_id(predictions: list[dict]) -> dict[str, dict]:
    return {str(p["id"]): p for p in predictions}


def _build_pairs(
    run_a_preds: dict[str, dict],
    run_b_preds: dict[str, dict],
    direction: str,
    per_direction: int,
    seed: int,
) -> list[dict]:
    """Select differing, usable A/B pairs for one direction, with per-row blinding."""
    rng = random.Random(f"{seed}:{direction}")
    candidates: list[dict] = []
    for rid, pa in run_a_preds.items():
        if pa.get("direction") != direction:
            continue
        pb = run_b_preds.get(rid)
        if pb is None or pb.get("direction") != direction:
            continue
        source = str(pa.get("source", ""))
        reference = str(pa.get("reference", ""))
        hyp_a = str(pa.get("hypothesis", ""))
        hyp_b = str(pb.get("hypothesis", ""))
        if not _usable(source, reference, hyp_a, hyp_b):
            continue
        if _normalized(hyp_a) == _normalized(hyp_b):
            continue  # identical output: no preference to express
        candidates.append(
            {
                "id": rid,
                "direction": direction,
                "source": source,
                "reference": reference,
                "run_a_hyp": hyp_a,
                "run_b_hyp": hyp_b,
            }
        )

    rng.shuffle(candidates)
    selected = candidates[: per_direction]
    selected.sort(key=lambda c: c["id"])

    rows: list[dict] = []
    for i, c in enumerate(selected, start=1):
        a_is_run_a = rng.random() < 0.5  # randomize which system is shown as "A"
        option_a = c["run_a_hyp"] if a_is_run_a else c["run_b_hyp"]
        option_b = c["run_b_hyp"] if a_is_run_a else c["run_a_hyp"]
        rows.append(
            {
                "N": i,
                "id": c["id"],
                "direction": direction,
                "source": c["source"],
                "reference": c["reference"],
                "option_a": option_a,
                "option_b": option_b,
                "A_system": "run_a" if a_is_run_a else "run_b",
                "B_system": "run_b" if a_is_run_a else "run_a",
            }
        )
    return rows


def _speaker_df(rows: list[dict]) -> pd.DataFrame:
    out = pd.DataFrame(
        {
            "N": [r["N"] for r in rows],
            "Texto fuente": [r["source"] for r in rows],
            "Referencia": [r["reference"] for r in rows],
            "Opción A": [r["option_a"] for r in rows],
            "Opción B": [r["option_b"] for r in rows],
            "Más cercana (A/B/Empate)": "",
            "Comentario": "",
        }
    )
    return out[SPEAKER_COLUMNS]


def _readme_df(run_a: str, run_b: str) -> pd.DataFrame:
    lines = [
        ("Objetivo", "Elegir, para cada fila, cuál de las dos opciones (A o B) está más cerca de la referencia."),
        ("Cómo revisar", "Lea la referencia y compare la Opción A y la Opción B. Marque A, B o Empate en la columna correspondiente."),
        ("Empates", "Se permiten empates: si ambas opciones son igual de cercanas (o ambas malas), escriba Empate."),
        ("Ciego", "El orden A/B es aleatorio y oculta qué sistema produjo cada opción; no intente adivinarlo."),
        ("Comentario", "Opcional: explique brevemente su elección o señale errores."),
        ("Nota", "La correspondencia entre A/B y cada sistema se guarda aparte, en el archivo de clave."),
    ]
    return pd.DataFrame(lines, columns=["campo", "detalle"])


def _format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = {"A": 6, "B": 40, "C": 40, "D": 40, "E": 40, "F": 18, "G": 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _write_workbook(output: Path, readme_df: pd.DataFrame, sheets: dict[str, pd.DataFrame]) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        readme_df.to_excel(writer, sheet_name="instrucciones", index=False)
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)
    wb = load_workbook(output)
    for name in wb.sheetnames:
        if name == "instrucciones":
            continue
        _format_sheet(wb[name])
    wb["instrucciones"].column_dimensions["A"].width = 18
    wb["instrucciones"].column_dimensions["B"].width = 110
    for cell in wb["instrucciones"][1]:
        cell.font = Font(bold=True)
    wb.save(output)


def main() -> int:
    args = parse_args()
    nmt = resolve_paths(PROJECT_ROOT, args.variant)
    run_a_path = nmt.reports_reranking_dir / args.run_a / "test_predictions_reranked.jsonl"
    run_b_path = nmt.reports_reranking_dir / args.run_b / "test_predictions_reranked.jsonl"
    output = args.output or (nmt.reports_evaluation_dir / "pairwise_preference.xlsx")
    anon_key_path = args.anon_key or (nmt.reports_evaluation_dir / "pairwise_preference_anon_key.json")

    for path in (run_a_path, run_b_path):
        if not path.exists():
            print(f"[phase8d] missing predictions: {path}", file=sys.stderr)
            return 2

    run_a_preds = _index_by_id(_read_jsonl(run_a_path))
    run_b_preds = _index_by_id(_read_jsonl(run_b_path))

    sheets: dict[str, pd.DataFrame] = {}
    anon_rows: list[dict] = []
    for direction in args.directions:
        rows = _build_pairs(run_a_preds, run_b_preds, direction, int(args.per_direction), int(args.seed))
        if not rows:
            print(f"[phase8d] no differing pairs for {direction}", file=sys.stderr)
            continue
        sheets[direction] = _speaker_df(rows)
        for r in rows:
            anon_rows.append({k: r[k] for k in ("N", "id", "direction", "A_system", "B_system")})
        print(f"[phase8d] {direction}: {len(rows)} pairs")

    if not sheets:
        print("[phase8d] nothing to write", file=sys.stderr)
        return 2

    _write_workbook(output, _readme_df(args.run_a, args.run_b), sheets)

    anon_key = {
        "run_a": args.run_a,
        "run_b": args.run_b,
        "directions": list(sheets.keys()),
        "seed": int(args.seed),
        "ties_allowed": True,
        "timestamp_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
        "rows": anon_rows,
    }
    anon_key_path.parent.mkdir(parents=True, exist_ok=True)
    anon_key_path.write_text(json.dumps(anon_key, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"[phase8d] wrote {_display_path(output)}")
    print(f"[phase8d] wrote {_display_path(anon_key_path)} (keep private)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
