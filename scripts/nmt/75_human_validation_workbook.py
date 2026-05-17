"""Build the 100-row human-validation workbook for the XL NMT champion.

The workbook is intended for Shiwilu/Spanish reviewers. It samples 50 test
predictions per direction from the champion reranked XL run, stratified by
sentence-level chrF++ quality buckets.

Output:
    reports/05_nmt/evaluation_xl/human_validation_100.xlsx
"""
from __future__ import annotations

import argparse
import json
import random
import re
import sys
from pathlib import Path
from typing import Iterable

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd  # noqa: E402
import sacrebleu  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from scripts.nmt._paths import resolve_paths  # noqa: E402


DEFAULT_RUN = "nllb_bidi_lora_v2_1b_loraplus_xl"
BUCKET_TARGETS = {"good": 35, "medium": 12, "mildly_bad": 3}
INTERNAL_COLUMNS = [
    "id",
    "direction",
    "origin_source",
    "quality_bucket",
    "chrf_pp",
    "confidence",
    "confidence_score",
    "texto_fuente",
    "referencia",
    "traduccion_modelo",
]
SPEAKER_COLUMNS = [
    "N",
    "Texto fuente",
    "Traducción del modelo",
    "Sentido (1-5)",
    "Naturalidad (1-5)",
    "Decisión",
    "Comentarios",
]


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--seed", type=int, default=2026)
    p.add_argument("--run", default=DEFAULT_RUN)
    p.add_argument(
        "--predictions",
        type=Path,
        default=None,
        help="Default: reports/05_nmt/reranking_xl/<run>/test_predictions_reranked.jsonl",
    )
    p.add_argument(
        "--test-csv",
        type=Path,
        default=None,
        help="Default: data/processed/06_nmt_filtered_xl/test.csv",
    )
    p.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Default: reports/05_nmt/evaluation_xl/human_validation_100.xlsx",
    )
    return p.parse_args()


def _read_jsonl(path: Path) -> list[dict]:
    rows: list[dict] = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def _display_path(path: Path) -> str:
    path = path.resolve()
    try:
        return str(path.relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path)


def _sentence_chrf_pp(hypothesis: str, reference: str) -> float:
    return float(
        sacrebleu.sentence_chrf(
            hypothesis,
            [reference],
            char_order=6,
            word_order=2,
            beta=2,
        ).score
    )


def _normalized(text: str) -> str:
    return re.sub(r"\s+", " ", str(text).strip()).casefold()


def _has_broken_encoding(text: str) -> bool:
    markers = ("�", "Ã", "Â", "â€™", "â€œ", "â€")
    return any(marker in str(text) for marker in markers)


def _has_severe_repetition(text: str) -> bool:
    tokens = re.findall(r"\w+(?:'\w+)?", str(text).casefold(), flags=re.UNICODE)
    if len(tokens) < 8:
        return False
    unique_ratio = len(set(tokens)) / len(tokens)
    most_common_ratio = max(tokens.count(t) for t in set(tokens)) / len(tokens)
    if unique_ratio <= 0.30 or most_common_ratio >= 0.45:
        return True

    for n in (2, 3):
        ngrams = [" ".join(tokens[i : i + n]) for i in range(len(tokens) - n + 1)]
        if ngrams and max(ngrams.count(g) for g in set(ngrams)) >= 4:
            return True
    return False


def _is_usable(row: dict) -> bool:
    fields = [row.get("texto_fuente", ""), row.get("referencia", ""), row.get("traduccion_modelo", "")]
    if any(not str(value).strip() for value in fields):
        return False
    if any(_has_broken_encoding(str(value)) for value in fields):
        return False
    return not _has_severe_repetition(str(row.get("traduccion_modelo", "")))


def _quality_bucket(hypothesis: str, reference: str, chrf_pp: float) -> str | None:
    if _normalized(hypothesis) == _normalized(reference) or chrf_pp >= 80:
        return "good"
    if 55 <= chrf_pp < 80:
        return "medium"
    if 35 <= chrf_pp < 55:
        return "mildly_bad"
    return None


def _sample_bucket(df: pd.DataFrame, bucket: str, n: int, rng: random.Random) -> pd.DataFrame:
    sub = df[df["quality_bucket"] == bucket].copy()
    if len(sub) < n:
        raise ValueError(f"Not enough {bucket!r} rows: need {n}, found {len(sub)}")

    if bucket == "good":
        exact = sub[sub["is_exact_match"]].copy()
        near = sub[~sub["is_exact_match"]].copy()
        parts: list[pd.DataFrame] = []
        if len(exact) >= n:
            return exact.sample(n=n, random_state=rng.randint(0, 1_000_000)).sort_values("id")
        if not exact.empty:
            parts.append(exact)
        needed = n - len(exact)
        if needed > 0:
            parts.append(near.sample(n=needed, random_state=rng.randint(0, 1_000_000)))
        return pd.concat(parts, ignore_index=True).sort_values(["is_exact_match", "chrf_pp", "id"], ascending=[False, False, True])

    return sub.sample(n=n, random_state=rng.randint(0, 1_000_000)).sort_values(["chrf_pp", "id"], ascending=[False, True])


def _sample_direction(df: pd.DataFrame, direction: str, seed: int) -> pd.DataFrame:
    rng = random.Random(f"{seed}:{direction}")
    sub = df[(df["direction"] == direction) & df.apply(lambda r: _is_usable(r.to_dict()), axis=1)].copy()
    pieces = [_sample_bucket(sub, bucket, n, rng) for bucket, n in BUCKET_TARGETS.items()]
    out = pd.concat(pieces, ignore_index=True)
    out["_bucket_order"] = out["quality_bucket"].map({b: i for i, b in enumerate(BUCKET_TARGETS)})
    out = out.sort_values(["_bucket_order", "chrf_pp", "id"], ascending=[True, False, True])
    return out[INTERNAL_COLUMNS].reset_index(drop=True)


def _build_rows(predictions_path: Path, test_csv: Path) -> pd.DataFrame:
    test_df = pd.read_csv(test_csv, encoding="utf-8-sig")
    test_by_id = test_df.set_index("id").to_dict(orient="index")
    rows: list[dict] = []

    for pred in _read_jsonl(predictions_path):
        rid = pred["id"]
        meta = test_by_id.get(rid, {})
        source = str(meta.get("source", pred.get("source", "")))
        reference = str(meta.get("target", pred.get("reference", "")))
        hypothesis = str(pred.get("hypothesis", ""))
        chrf_pp = _sentence_chrf_pp(hypothesis, reference)
        bucket = _quality_bucket(hypothesis, reference, chrf_pp)
        if bucket is None:
            continue

        rows.append(
            {
                "id": rid,
                "direction": pred.get("direction", ""),
                "origin_source": meta.get("origin_source", "unknown"),
                "quality_bucket": bucket,
                "chrf_pp": round(chrf_pp, 2),
                "confidence": pred.get("confidence", ""),
                "confidence_score": pred.get("confidence_score", ""),
                "texto_fuente": source,
                "referencia": reference,
                "traduccion_modelo": hypothesis,
                "is_exact_match": _normalized(hypothesis) == _normalized(reference),
            }
        )
    return pd.DataFrame(rows)


def _readme_rows(predictions_path: Path) -> pd.DataFrame:
    lines = [
        ("Objetivo", "Revisar 100 traducciones del modelo: 50 de shiwilu a castellano y 50 de castellano a shiwilu."),
        ("Como revisar", "Lea el texto fuente y la traduccion del modelo. No hay una referencia visible: juzgue si la traduccion comunica bien el sentido."),
        ("Sentido (1-5)", "5 = conserva bien el significado; 1 = cambia el sentido o falta informacion importante."),
        ("Naturalidad (1-5)", "5 = suena natural en la lengua de llegada; 1 = suena raro, forzado o dificil de entender."),
        ("Decisión", "Escriba aceptar, corregir o rechazar."),
        ("Comentarios", "Escriba cualquier duda, correccion sugerida o explicacion breve."),
        ("Nota", "No necesita revisar ningun dato tecnico. La seleccion ya fue preparada automaticamente."),
    ]
    return pd.DataFrame(lines, columns=["campo", "detalle"])


def _speaker_sheet(df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame(
        {
            "N": range(1, len(df) + 1),
            "Texto fuente": df["texto_fuente"],
            "Traducción del modelo": df["traduccion_modelo"],
            "Sentido (1-5)": "",
            "Naturalidad (1-5)": "",
            "Decisión": "",
            "Comentarios": "",
        }
    )
    return out[SPEAKER_COLUMNS]


def _autosize_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = {
        "A": 6,
        "B": 44,
        "C": 44,
        "D": 13,
        "E": 12,
        "F": 12,
        "G": 36,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _write_workbook(output: Path, readme_df: pd.DataFrame, sheets: dict[str, pd.DataFrame]) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        readme_df.to_excel(writer, sheet_name="instrucciones", index=False)
        for name, df in sheets.items():
            _speaker_sheet(df).to_excel(writer, sheet_name=name, index=False)

    wb = load_workbook(output)
    for ws in wb.worksheets:
        _autosize_sheet(ws)
    wb["instrucciones"].column_dimensions["A"].width = 24
    wb["instrucciones"].column_dimensions["B"].width = 110
    wb.save(output)


def _validate_workbook(output: Path) -> None:
    required = set(SPEAKER_COLUMNS)
    wb = load_workbook(output)
    expected_sheets = {"instrucciones", "shw2spa", "spa2shw"}
    if set(wb.sheetnames) != expected_sheets:
        raise AssertionError(f"Unexpected sheets: {wb.sheetnames}")

    for direction in ("shw2spa", "spa2shw"):
        ws = wb[direction]
        headers = [cell.value for cell in ws[1]]
        missing = required - set(headers)
        if missing:
            raise AssertionError(f"{direction} missing columns: {sorted(missing)}")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if len(rows) != 50:
            raise AssertionError(f"{direction} has {len(rows)} data rows, expected 50")

        idx = {name: headers.index(name) for name in headers}
        for row in rows:
            for col in ("Texto fuente", "Traducción del modelo"):
                if not str(row[idx[col]] or "").strip():
                    raise AssertionError(f"{direction} has an empty {col}")


def _validate_sampled_sheets(sheets: dict[str, pd.DataFrame]) -> None:
    for direction, df in sheets.items():
        if len(df) != 50:
            raise AssertionError(f"{direction} has {len(df)} sampled rows, expected 50")
        counts = df["quality_bucket"].value_counts().to_dict()
        if counts != BUCKET_TARGETS:
            raise AssertionError(f"{direction} bucket counts {counts}, expected {BUCKET_TARGETS}")
        if (df.loc[df["quality_bucket"] == "mildly_bad", "chrf_pp"] < 35).any():
            raise AssertionError(f"{direction} has mildly_bad row with chrF++ < 35")
        for col in ("texto_fuente", "referencia", "traduccion_modelo"):
            if df[col].astype(str).str.strip().eq("").any():
                raise AssertionError(f"{direction} has an empty {col}")


def _summarize_counts(sheets: Iterable[tuple[str, pd.DataFrame]]) -> str:
    chunks = []
    for direction, df in sheets:
        counts = df["quality_bucket"].value_counts().reindex(BUCKET_TARGETS.keys()).fillna(0).astype(int)
        chunks.append(f"{direction}: " + ", ".join(f"{k}={v}" for k, v in counts.items()))
    return "; ".join(chunks)


def main() -> int:
    args = parse_args()
    nmt = resolve_paths(PROJECT_ROOT, "xl")
    predictions_path = args.predictions or (nmt.reports_reranking_dir / args.run / "test_predictions_reranked.jsonl")
    test_csv = args.test_csv or (nmt.filtered_dir / "test.csv")
    output = args.output or (nmt.reports_evaluation_dir / "human_validation_100.xlsx")

    if not predictions_path.exists():
        raise FileNotFoundError(predictions_path)
    if not test_csv.exists():
        raise FileNotFoundError(test_csv)

    scored = _build_rows(predictions_path, test_csv)
    sheets = {
        "shw2spa": _sample_direction(scored, "shw2spa", args.seed),
        "spa2shw": _sample_direction(scored, "spa2shw", args.seed),
    }
    _validate_sampled_sheets(sheets)
    _write_workbook(output, _readme_rows(predictions_path), sheets)
    _validate_workbook(output)

    print(f"[phase8c] wrote {_display_path(output)}")
    print(f"[phase8c] validated workbook: {_summarize_counts(sheets.items())}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
